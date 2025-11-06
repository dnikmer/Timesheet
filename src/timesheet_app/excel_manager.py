"""Вспомогательные функции для работы с Excel.

Содержит:
- константы имён листов;
- загрузку справочников (проекты и виды работ);
- добавление записи о затраченном времени;
- создание шаблонной книги с нужными листами и заголовками.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, List, Tuple

from openpyxl import Workbook, load_workbook


# Имена листов в книге Excel
REFERENCE_SHEET = "Справочник"
TIMESHEET_SHEET = "Учет времени"
WORKDAY_SHEET = "Учет рабочего времени"


class ExcelStructureError(RuntimeError):
    """Структура книги Excel не соответствует ожиданиям."""


def _normalise(values: Iterable[str | None]) -> List[str]:
    """Очистка и нормализация значений (удаляем пустые, дубликаты)."""

    items: List[str] = []
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text and text not in items:
            items.append(text)
    return items


def load_reference_data(path: Path | str) -> Tuple[List[str], List[str]]:
    """Прочитать лист справочника и вернуть два списка: проекты и виды работ."""

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)

    if REFERENCE_SHEET not in workbook:
        raise ExcelStructureError(
            f"Workbook must contain sheet '{REFERENCE_SHEET}'. Found: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[REFERENCE_SHEET]

    projects: List[str] = []
    work_types: List[str] = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Пропускаем возможную строку заголовков
        if idx == 1:
            continue
        project, work_type, *_ = row + (None, None)
        projects.append(project)
        work_types.append(work_type)

    return _normalise(projects), _normalise(work_types)


def append_time_entry(
    path: Path | str,
    *,
    project: str,
    work_type: str,
    elapsed_seconds: float,
    finished_at: datetime | None = None,
) -> None:
    """Добавить строку на лист учёта времени (дата, проект, вид работ, длительность).

    В отличие от простого `sheet.append`, мы ищем первую по-настоящему пустую
    строку, игнорируя форматирование (цвета, границы) и удалённые строки.
    """

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    workbook = load_workbook(workbook_path)

    if TIMESHEET_SHEET not in workbook:
        raise ExcelStructureError(
            f"Workbook must contain sheet '{TIMESHEET_SHEET}'. Found: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[TIMESHEET_SHEET]
    timestamp = finished_at or datetime.now()

    duration = timedelta(seconds=elapsed_seconds)

    # Поиск первой полностью пустой строки, начиная со 2-й (после заголовков)
    def is_empty_row(r: int) -> bool:
        return (
            sheet.cell(row=r, column=1).value is None
            and sheet.cell(row=r, column=2).value is None
            and sheet.cell(row=r, column=3).value is None
            and sheet.cell(row=r, column=4).value is None
        )

    first = 2
    last = sheet.max_row
    target_row = None
    for r in range(first, last + 1):
        if is_empty_row(r):
            target_row = r
            break
    if target_row is None:
        target_row = last + 1

    # Записываем значения по ячейкам — так мы не зависим от sheet.append
    c_date = sheet.cell(row=target_row, column=1)
    c_date.value = timestamp.date()
    # Дата в формате ДД.ММ.ГГГГ
    try:
        c_date.number_format = "DD.MM.YYYY"
    except Exception:
        pass
    sheet.cell(row=target_row, column=2).value = project
    sheet.cell(row=target_row, column=3).value = work_type
    cell_duration = sheet.cell(row=target_row, column=4)
    cell_duration.value = duration
    # Красивый формат времени часов:минуты:секунды
    try:
        cell_duration.number_format = "[h]:mm:ss"
    except Exception:
        pass

    workbook.save(workbook_path)


def _first_empty_row(sheet, start_row: int, last_col: int) -> int:
    """Найти первую полностью пустую строку (значения None) начиная с `start_row`.

    Учитывается только содержимое ячеек, любые стили/границы игнорируются.
    Возвращает номер строки (>= start_row).
    """

    def row_empty(r: int) -> bool:
        for col in range(1, last_col + 1):
            if sheet.cell(row=r, column=col).value is not None:
                return False
        return True

    last = sheet.max_row
    for r in range(start_row, last + 1):
        if row_empty(r):
            return r
    return last + 1


def workday_start(path: Path | str) -> tuple[str, str]:
    """Записать текущую дату и время начала в лист "Учет рабочего времени".

    Возвращает пару строк для сообщений: (дата ДД.ММ.ГГГГ, время ЧЧ:ММ).
    """

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    wb = load_workbook(workbook_path)

    if WORKDAY_SHEET not in wb:
        # Создадим лист при первом использовании
        ws = wb.create_sheet(WORKDAY_SHEET)
        ws.append(["Дата", "Время начала", "Время окончания", "Длительность"])
    else:
        ws = wb[WORKDAY_SHEET]

    target_row = _first_empty_row(ws, start_row=2, last_col=4)
    now = datetime.now()
    date_str = now.strftime("%d.%m.%Y")
    time_str = now.strftime("%H:%M")

    c_date = ws.cell(row=target_row, column=1)
    c_date.value = now.date()
    try:
        c_date.number_format = "DD.MM.YYYY"
    except Exception:
        pass

    c_start = ws.cell(row=target_row, column=2)
    c_start.value = now.time()
    try:
        c_start.number_format = "HH:MM"
    except Exception:
        pass

    wb.save(workbook_path)
    return date_str, time_str


def workday_end(path: Path | str) -> str:
    """Записать время окончания и длительность в лист "Учет рабочего времени".

    Ищет последнюю строку, где заполнены дата/время начала, но пусто время окончания.
    Возвращает строку длительности в формате ЧЧ:ММ для сообщений.
    """

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    wb = load_workbook(workbook_path)
    if WORKDAY_SHEET not in wb:
        raise ExcelStructureError(f"Workbook must contain sheet '{WORKDAY_SHEET}'.")

    ws = wb[WORKDAY_SHEET]

    # Ищем последнюю незавершённую запись
    target_row = None
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=1).value is not None and ws.cell(row=r, column=2).value is not None and ws.cell(row=r, column=3).value is None:
            target_row = r
            break

    if target_row is None:
        raise ExcelStructureError("Не найдено незавершённое начало рабочего дня.")

    now = datetime.now()
    # Записываем время окончания
    c_end = ws.cell(row=target_row, column=3)
    c_end.value = now.time()
    try:
        c_end.number_format = "HH:MM"
    except Exception:
        pass

    # Рассчитываем длительность
    start_cell = ws.cell(row=target_row, column=2)
    date_cell = ws.cell(row=target_row, column=1)

    # Приводим к datetime для вычисления
    start_time = start_cell.value
    start_date = date_cell.value
    if isinstance(start_date, datetime):
        start_date = start_date.date()
    if isinstance(start_time, datetime):
        start_time = start_time.time()

    start_dt = datetime.combine(start_date, start_time)
    duration = now - start_dt
    # Округляем до минут
    minutes = int(duration.total_seconds() // 60)
    hours, mins = divmod(minutes, 60)
    dur_str = f"{hours:02d}:{mins:02d}"

    c_dur = ws.cell(row=target_row, column=4)
    # Для Excel пишем как timedelta, а форматируем как ч:мм
    c_dur.value = timedelta(hours=hours, minutes=mins)
    try:
        c_dur.number_format = "[h]:mm"
    except Exception:
        pass

    wb.save(workbook_path)
    return dur_str


def create_template(path: Path | str) -> None:
    """Создать пустую книгу Excel с нужными листами и заголовками."""

    workbook_path = Path(path)
    wb = Workbook()
    # Удалим дефолтный лист, чтобы контролировать порядок
    default = wb.active
    wb.remove(default)

    # Лист справочника
    ws_ref = wb.create_sheet(REFERENCE_SHEET)
    ws_ref.append(["Проект", "Вид работ"])  # заголовки

    # Лист учёта времени (суммарные записи)
    ws_ts = wb.create_sheet(TIMESHEET_SHEET)
    ws_ts.append(["Дата", "Проект", "Вид работ", "Длительность"])  # заголовки

    # Лист учёта рабочего дня (начала/окончания)
    ws_wd = wb.create_sheet(WORKDAY_SHEET)
    ws_wd.append(["Дата", "Время начала", "Время окончания", "Длительность"])  # заголовки

    wb.save(workbook_path)
